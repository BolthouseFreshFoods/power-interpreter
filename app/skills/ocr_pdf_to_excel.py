"""Skill: OCR PDF to Excel — Automated document processing.

Multi-step workflow that:
1. Acquires a PDF (from sandbox or OneDrive)
2. Converts PDF pages to images (pdf2image)
3. Runs OCR on each page (pytesseract)
4. Writes structured results to a formatted Excel workbook
5. Returns download URL for the output file

Designed for Cassidi's handwritten/scanned PDF workflow.
Reduces 4-6 agent tool calls to 1 skill call.
"""

import logging
import re
import json
import asyncio

logger = logging.getLogger(__name__)

# Download URL pattern for Power Interpreter on Railway
_DL_URL_PATTERN = (
    r'https://power-interpreter-production-6396'
    r'\.up\.railway\.app/dl/[a-f0-9-]+'
)

# UUID pattern for session ID extraction
_UUID_PATTERN = (
    r'[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-'
    r'[0-9a-f]{4}-[0-9a-f]{12}'
)


async def _ensure_session(engine, session_id: str) -> str:
    """Ensure a session exists, creating one if needed."""
    if session_id and session_id.strip():
        return session_id.strip()

    result = await engine.call_tool("create_session")
    try:
        parsed = json.loads(result)
        sid = parsed.get("session_id", parsed.get("id", ""))
        if sid:
            return sid
    except (json.JSONDecodeError, AttributeError):
        pass

    match = re.search(_UUID_PATTERN, result)
    if match:
        return match.group()

    raise RuntimeError(f"Could not create session: {result[:200]}")


async def _get_download_url(engine, session_id: str, filename: str) -> str:
    """Get download URL for a file, with retry for list_files unreliability."""
    for attempt in range(3):
        files_result = await engine.call_tool("list_files", session_id=session_id)
        if filename in files_result:
            match = re.search(_DL_URL_PATTERN, files_result)
            if match:
                return match.group()
        if attempt < 2:
            logger.info(f"list_files retry {attempt + 1}/3 for {filename}")
            await asyncio.sleep(1.5)
    return ""


async def execute_ocr_pdf_to_excel(
    engine,
    source_type: str = "local",
    file_path: str = "",
    user_id: str = "",
    item_id: str = "",
    output_filename: str = "ocr_output.xlsx",
    pages: str = "all",
    session_id: str = "",
) -> str:
    """OCR a PDF and extract text into a structured Excel workbook.

    Args:
        engine: SkillEngine instance (injected)
        source_type: 'local' (sandbox path) or 'onedrive' (download first)
        file_path: Path to PDF in sandbox (for local source)
        user_id: Microsoft 365 email (for OneDrive source)
        item_id: OneDrive file ID (for OneDrive source)
        output_filename: Name for output Excel file
        pages: 'all' or page spec like '1-3' or '1,3,5'
        session_id: Session ID (auto-created if empty)

    Returns:
        Status message with download URL or error details
    """
    steps = []

    try:
        # -- Validate inputs --
        if source_type == "onedrive":
            if not user_id:
                return "Error: user_id is required for OneDrive source."
            if not item_id:
                return "Error: item_id is required for OneDrive source."
        elif source_type == "local":
            if not file_path:
                return "Error: file_path is required for local source."
        else:
            return (
                f"Error: Unknown source_type '{source_type}'. "
                f"Use 'local' or 'onedrive'."
            )

        # -- Step 1: Session --
        session_id = await _ensure_session(engine, session_id)
        steps.append(f"Session ready ({session_id[:8]}...)")

        # -- Step 2: Acquire PDF --
        pdf_path = file_path
        if source_type == "onedrive":
            logger.info(
                f"skill_ocr_pdf_to_excel: downloading {item_id} "
                f"for {user_id}"
            )
            dl_result = await engine.call_tool(
                "onedrive",
                action="download",
                user_id=user_id,
                item_id=item_id,
                session_id=session_id,
            )
            if "error" in dl_result.lower() and "saved" not in dl_result.lower():
                return (
                    f"Step 2 failed: OneDrive download error.\n\n"
                    f"{dl_result[:500]}"
                )
            # Extract sandbox_path
            try:
                dl_parsed = json.loads(dl_result)
                pdf_path = dl_parsed.get("sandbox_path", "")
            except (json.JSONDecodeError, AttributeError):
                pass
            if not pdf_path:
                path_match = re.search(
                    r'/app/sandbox_data/[^\s"]+\.pdf',
                    dl_result, re.IGNORECASE,
                )
                if path_match:
                    pdf_path = path_match.group()
            if not pdf_path:
                return (
                    f"Step 2 failed: Could not find sandbox path "
                    f"in download result.\n\n{dl_result[:500]}"
                )
            steps.append(f"PDF downloaded to {pdf_path}")
        else:
            steps.append(f"Using local PDF: {pdf_path}")

        # -- Step 3: OCR + Excel generation --
        safe_pdf = pdf_path.replace("'", "\\'")
        safe_out = (
            "/app/sandbox_data/" + session_id + "/" + output_filename
        ).replace("'", "\\'")

        # Build page selection code
        page_block = ""
        if pages != "all":
            safe_pages = pages.replace("'", "\\'")
            page_block = (
                "selected = []\n"
                "for part in '" + safe_pages + "'.split(','):\n"
                "    part = part.strip()\n"
                "    if '-' in part:\n"
                "        s, e = part.split('-', 1)\n"
                "        selected.extend(range(int(s)-1, int(e)))\n"
                "    else:\n"
                "        selected.append(int(part)-1)\n"
                "images = [images[i] for i in selected if i < len(images)]\n"
                "print(f'Selected {len(images)} pages from spec')\n"
            )

        ocr_code = (
            "from pdf2image import convert_from_path\n"
            "import pytesseract\n"
            "from openpyxl import Workbook\n"
            "from openpyxl.styles import Font, PatternFill\n"
            "\n"
            "print('Converting PDF to images...')\n"
            "images = convert_from_path('" + safe_pdf + "')\n"
            "print(f'Converted {len(images)} pages')\n"
            "\n"
            + page_block +
            "\n"
            "ocr_results = []\n"
            "for i, img in enumerate(images):\n"
            "    print(f'OCR page {i+1}/{len(images)}...')\n"
            "    text = pytesseract.image_to_string(img)\n"
            "    lines = [l.strip() for l in text.split('\\n') if l.strip()]\n"
            "    ocr_results.append({\n"
            "        'page': i + 1, 'text': text,\n"
            "        'lines': lines, 'line_count': len(lines),\n"
            "        'char_count': len(text),\n"
            "    })\n"
            "\n"
            "# -- Excel workbook --\n"
            "wb = Workbook()\n"
            "ws = wb.active\n"
            "ws.title = 'Summary'\n"
            "hf = Font(bold=True, color='FFFFFF')\n"
            "hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')\n"
            "\n"
            "ws.append(['OCR Extraction Summary'])\n"
            "ws['A1'].font = Font(bold=True, size=14)\n"
            "ws.append([])\n"
            "ws.append(['Total Pages', len(ocr_results)])\n"
            "ws.append(['Total Lines', sum(r['line_count'] for r in ocr_results)])\n"
            "ws.append(['Total Characters', sum(r['char_count'] for r in ocr_results)])\n"
            "ws.append([])\n"
            "ws.append(['Page', 'Lines', 'Characters', 'Preview'])\n"
            "for c in ws[ws.max_row]:\n"
            "    c.font = hf\n"
            "    c.fill = hfill\n"
            "for r in ocr_results:\n"
            "    preview = r['lines'][0][:80] if r['lines'] else '(empty)'\n"
            "    ws.append([r['page'], r['line_count'], r['char_count'], preview])\n"
            "\n"
            "# Raw text sheet\n"
            "ws2 = wb.create_sheet('Raw Text')\n"
            "ws2.append(['Page', 'Line', 'Text'])\n"
            "for c in ws2[1]:\n"
            "    c.font = hf\n"
            "    c.fill = hfill\n"
            "for r in ocr_results:\n"
            "    for j, line in enumerate(r['lines']):\n"
            "        ws2.append([r['page'], j + 1, line])\n"
            "\n"
            "# Per-page sheets (max 10)\n"
            "for r in ocr_results[:10]:\n"
            "    sn = f\"Page {r['page']}\"\n"
            "    wp = wb.create_sheet(sn[:31])\n"
            "    wp.append(['Line', 'Text'])\n"
            "    for c in wp[1]:\n"
            "        c.font = hf\n"
            "        c.fill = hfill\n"
            "    for j, line in enumerate(r['lines']):\n"
            "        wp.append([j + 1, line])\n"
            "    wp.column_dimensions['B'].width = 80\n"
            "\n"
            "wb.save('" + safe_out + "')\n"
            "print(f'\\nSaved OCR results to: " + safe_out + "')\n"
            "print(f'Total: {sum(r[\"line_count\"] for r in ocr_results)}'"
            "      f' lines from {len(ocr_results)} pages')\n"
        )

        logger.info(
            f"skill_ocr_pdf_to_excel: executing OCR on {pdf_path}"
        )
        exec_result = await engine.call_tool(
            "execute_code",
            session_id=session_id,
            code=ocr_code,
        )

        if "error" in exec_result.lower() and "saved" not in exec_result.lower():
            return (
                f"Step 3 failed: OCR execution error.\n\n"
                f"{exec_result[:2000]}"
            )
        steps.append("OCR completed + Excel generated")

        # -- Step 4: Download URL --
        download_url = await _get_download_url(
            engine, session_id, output_filename
        )
        if download_url:
            steps.append("Download URL retrieved")
        else:
            steps.append("File created but URL not captured")

        # -- Report --
        report = [
            "## OCR PDF to Excel \u2014 Complete",
            "",
            f"**Source:** `{pdf_path}`",
            f"**Output:** `{output_filename}`",
        ]
        if download_url:
            report.append(f"**Download:** {download_url}")
        report.append("")
        report.append("### Steps:")
        for s in steps:
            report.append(f"- {s}")
        report.append("")
        report.append("### OCR Output:")
        report.append(exec_result[:2000] if exec_result else "(no output)")
        return "\n".join(report)

    except Exception as e:
        logger.error(f"skill_ocr_pdf_to_excel: {e}", exc_info=True)
        lines = [
            "## OCR PDF to Excel \u2014 Failed",
            f"**Error:** {e}",
            "",
            "### Steps completed:",
        ]
        for s in (steps or ["(none)"]):
            lines.append(f"- {s}")
        return "\n".join(lines)


SKILL_DEFINITION = {
    "name": "skill_ocr_pdf_to_excel",
    "description": (
        "OCR a PDF file and extract text into a structured Excel workbook. "
        "Supports local sandbox files or OneDrive downloads. Converts PDF "
        "pages to images, runs OCR (pytesseract), and writes structured "
        "results with Summary, Raw Text, and per-page sheets. "
        "For local: pass source_type='local' and file_path. "
        "For OneDrive: pass source_type='onedrive' with user_id and item_id."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "source_type": {
                "type": "string",
                "description": "'local' (sandbox) or 'onedrive'",
                "default": "local",
                "enum": ["local", "onedrive"],
            },
            "file_path": {
                "type": "string",
                "description": "Path to PDF in sandbox (for local)",
                "default": "",
            },
            "user_id": {
                "type": "string",
                "description": "Microsoft 365 email (for OneDrive)",
                "default": "",
            },
            "item_id": {
                "type": "string",
                "description": "OneDrive file ID (for OneDrive)",
                "default": "",
            },
            "output_filename": {
                "type": "string",
                "description": "Output Excel filename",
                "default": "ocr_output.xlsx",
            },
            "pages": {
                "type": "string",
                "description": "'all' or page spec like '1-3' or '1,3,5'",
                "default": "all",
            },
            "session_id": {
                "type": "string",
                "description": "Session ID (auto-created if empty)",
                "default": "",
            },
        },
        "required": [],
    },
    "tools": ["onedrive", "execute_code", "list_files", "create_session"],
    "execute": execute_ocr_pdf_to_excel,
}
