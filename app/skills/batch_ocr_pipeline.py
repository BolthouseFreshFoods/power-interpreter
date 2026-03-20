"""Skill: Batch OCR Pipeline — Process multiple PDFs from OneDrive.

Multi-step workflow that:
1. Lists files in an OneDrive folder
2. Filters to PDF files
3. Downloads each PDF to sandbox
4. Runs OCR on all PDFs (pytesseract + pdf2image)
5. Combines results into a single Excel workbook
   (one sheet per PDF + summary sheet)
6. Returns download URL for the combined output

Designed for Cassidi's batch PDF processing from /AI Trial/PDFs.
Reduces 40-60 agent tool calls to 1 skill call.
"""

import logging
import re
import json
import asyncio

logger = logging.getLogger(__name__)

_DL_URL_PATTERN = (
    r'https://power-interpreter-production-6396'
    r'\.up\.railway\.app/dl/[a-f0-9-]+'
)

_UUID_PATTERN = (
    r'[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-'
    r'[0-9a-f]{4}-[0-9a-f]{12}'
)


async def _ensure_session(engine, session_id: str) -> str:
    if session_id and session_id.strip():
        return session_id.strip()
    result = await engine.call_tool("create_session")
    try:
        parsed = json.loads(result)
        sid = parsed.get("session_id", parsed.get("id", ""))
        if sid:
            return sid
    except (json.JSONDecodeError, AttributeError):
        pass
    match = re.search(_UUID_PATTERN, result)
    if match:
        return match.group()
    raise RuntimeError(f"Could not create session: {result[:200]}")


async def _get_download_url(engine, session_id: str, filename: str) -> str:
    for attempt in range(3):
        files_result = await engine.call_tool("list_files", session_id=session_id)
        if filename in files_result:
            match = re.search(_DL_URL_PATTERN, files_result)
            if match:
                return match.group()
        if attempt < 2:
            await asyncio.sleep(1.5)
    return ""


async def execute_batch_ocr_pipeline(
    engine,
    user_id: str = "",
    source_folder: str = "/",
    file_filter: str = ".pdf",
    output_filename: str = "batch_ocr_results.xlsx",
    max_files: int = 10,
    session_id: str = "",
) -> str:
    """OCR multiple PDFs from OneDrive into a combined Excel workbook.

    Args:
        engine: SkillEngine instance (injected)
        user_id: Microsoft 365 email (REQUIRED)
        source_folder: OneDrive folder path containing PDFs
        file_filter: Extension filter (default '.pdf')
        output_filename: Combined output Excel filename
        max_files: Safety cap on number of PDFs (default 10)
        session_id: Session ID (auto-created if empty)

    Returns:
        Status message with download URL or error details
    """
    steps = []

    try:
        if not user_id:
            return "Error: user_id is REQUIRED for batch OCR pipeline."

        # -- Step 1: Session --
        session_id = await _ensure_session(engine, session_id)
        steps.append(f"Session ready ({session_id[:8]}...)")

        # -- Step 2: List OneDrive folder --
        logger.info(
            f"skill_batch_ocr: listing '{source_folder}' for {user_id}"
        )
        list_result = await engine.call_tool(
            "onedrive",
            action="list",
            user_id=user_id,
            path=source_folder,
        )
        if "error" in list_result.lower() and "count" not in list_result.lower():
            return (
                f"Step 2 failed: Could not list OneDrive folder "
                f"'{source_folder}'.\n\n{list_result[:500]}"
            )
        steps.append(f"Listed files in '{source_folder}'")

        # -- Step 3: Filter to PDFs and extract IDs --
        pdf_files = []
        try:
            parsed = json.loads(list_result)
            items = parsed.get("items", parsed.get("files", []))
            if isinstance(items, list):
                for item in items:
                    if isinstance(item, dict):
                        name = item.get("name", "")
                        item_id = item.get("id", "")
                        if (
                            name.lower().endswith(file_filter.lower())
                            and item_id
                        ):
                            pdf_files.append(
                                {"name": name, "id": item_id}
                            )
        except (json.JSONDecodeError, AttributeError) as e:
            logger.warning(f"Could not parse file listing: {e}")
            return (
                f"Step 3 failed: Could not parse file listing.\n\n"
                f"{list_result[:500]}"
            )

        if not pdf_files:
            return (
                f"No PDF files found in '{source_folder}' "
                f"matching filter '{file_filter}'.\n\n"
                f"Files found: {list_result[:500]}"
            )

        # Apply safety cap
        if len(pdf_files) > max_files:
            logger.info(
                f"Capping from {len(pdf_files)} to {max_files} PDFs"
            )
            pdf_files = pdf_files[:max_files]

        steps.append(
            f"Found {len(pdf_files)} PDFs to process"
        )

        # -- Step 4: Download PDFs --
        file_ids_str = ",".join(f["id"] for f in pdf_files)
        dl_result = await engine.call_tool(
            "onedrive",
            action="batch_download",
            user_id=user_id,
            file_ids=file_ids_str,
            session_id=session_id,
        )

        # Parse download results to get sandbox paths
        sandbox_paths = []
        try:
            dl_parsed = json.loads(dl_result)
            files_list = dl_parsed.get("files", [])
            for f in files_list:
                if f.get("status") == "ok":
                    sp = f.get("sandbox_path", "")
                    name = f.get("name", "")
                    if sp and name.lower().endswith(".pdf"):
                        sandbox_paths.append(
                            {"name": name, "path": sp}
                        )
        except (json.JSONDecodeError, AttributeError):
            # Try regex fallback
            for f in pdf_files:
                path_match = re.search(
                    r'/app/sandbox_data/[^\s"]*'
                    + re.escape(f["name"]),
                    dl_result,
                )
                if path_match:
                    sandbox_paths.append(
                        {"name": f["name"], "path": path_match.group()}
                    )

        if not sandbox_paths:
            return (
                f"Step 4 failed: No PDFs were downloaded "
                f"successfully.\n\n{dl_result[:500]}"
            )

        steps.append(
            f"Downloaded {len(sandbox_paths)}/{len(pdf_files)} PDFs"
        )

        # -- Step 5: OCR all PDFs + combine --
        # Build paths list for the sandbox code
        paths_json = json.dumps(sandbox_paths)
        safe_out = (
            "/app/sandbox_data/" + session_id + "/" + output_filename
        ).replace("'", "\\'")

        ocr_batch_code = (
            "from pdf2image import convert_from_path\n"
            "import pytesseract\n"
            "from openpyxl import Workbook\n"
            "from openpyxl.styles import Font, PatternFill\n"
            "import json\n"
            "\n"
            "pdf_files = json.loads('" + paths_json.replace("'", "\\'") + "')\n"
            "print(f'Processing {len(pdf_files)} PDFs...')\n"
            "\n"
            "all_results = []\n"
            "for idx, pf in enumerate(pdf_files):\n"
            "    name = pf['name']\n"
            "    path = pf['path']\n"
            "    print(f'\\n[{idx+1}/{len(pdf_files)}] {name}')\n"
            "    try:\n"
            "        images = convert_from_path(path)\n"
            "        print(f'  Converted to {len(images)} page(s)')\n"
            "        pages = []\n"
            "        for pi, img in enumerate(images):\n"
            "            text = pytesseract.image_to_string(img)\n"
            "            lines = [l.strip() for l in text.split('\\n') if l.strip()]\n"
            "            pages.append({\n"
            "                'page': pi + 1,\n"
            "                'lines': lines,\n"
            "                'line_count': len(lines),\n"
            "                'char_count': len(text),\n"
            "            })\n"
            "            print(f'  Page {pi+1}: {len(lines)} lines')\n"
            "        all_results.append({\n"
            "            'name': name,\n"
            "            'page_count': len(images),\n"
            "            'pages': pages,\n"
            "            'total_lines': sum(p['line_count'] for p in pages),\n"
            "        })\n"
            "    except Exception as e:\n"
            "        print(f'  ERROR: {e}')\n"
            "        all_results.append({\n"
            "            'name': name, 'page_count': 0,\n"
            "            'pages': [], 'total_lines': 0, 'error': str(e),\n"
            "        })\n"
            "\n"
            "# -- Build combined Excel --\n"
            "wb = Workbook()\n"
            "hf = Font(bold=True, color='FFFFFF')\n"
            "hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')\n"
            "\n"
            "# Summary sheet\n"
            "ws = wb.active\n"
            "ws.title = 'Summary'\n"
            "ws.append(['Batch OCR Summary'])\n"
            "ws['A1'].font = Font(bold=True, size=14)\n"
            "ws.append([])\n"
            "ws.append(['Total PDFs', len(all_results)])\n"
            "ws.append(['Total Pages', sum(r['page_count'] for r in all_results)])\n"
            "ws.append(['Total Lines', sum(r['total_lines'] for r in all_results)])\n"
            "ok = [r for r in all_results if 'error' not in r]\n"
            "fail = [r for r in all_results if 'error' in r]\n"
            "ws.append(['Successful', len(ok)])\n"
            "ws.append(['Failed', len(fail)])\n"
            "ws.append([])\n"
            "ws.append(['PDF Name', 'Pages', 'Lines', 'Status'])\n"
            "for c in ws[ws.max_row]:\n"
            "    c.font = hf\n"
            "    c.fill = hfill\n"
            "for r in all_results:\n"
            "    status = 'OK' if 'error' not in r else f\"ERROR: {r['error'][:50]}\"\n"
            "    ws.append([r['name'], r['page_count'], r['total_lines'], status])\n"
            "\n"
            "# Per-PDF sheets (max 15 to avoid huge workbooks)\n"
            "for r in all_results[:15]:\n"
            "    if 'error' in r or not r['pages']:\n"
            "        continue\n"
            "    sn = r['name'][:28].replace('.pdf', '')\n"
            "    ws_p = wb.create_sheet(sn)\n"
            "    ws_p.append(['Page', 'Line', 'Text'])\n"
            "    for c in ws_p[1]:\n"
            "        c.font = hf\n"
            "        c.fill = hfill\n"
            "    for pg in r['pages']:\n"
            "        for j, line in enumerate(pg['lines']):\n"
            "            ws_p.append([pg['page'], j + 1, line])\n"
            "    ws_p.column_dimensions['C'].width = 80\n"
            "\n"
            "wb.save('" + safe_out + "')\n"
            "print(f'\\nCombined workbook saved to: " + safe_out + "')\n"
            "print(f'Processed {len(ok)} PDFs successfully, {len(fail)} failed')\n"
            "print(f'Total: {sum(r[\"total_lines\"] for r in all_results)} lines extracted')\n"
        )

        logger.info(
            f"skill_batch_ocr: executing OCR on {len(sandbox_paths)} PDFs"
        )
        exec_result = await engine.call_tool(
            "execute_code",
            session_id=session_id,
            code=ocr_batch_code,
        )

        if "error" in exec_result.lower() and "saved" not in exec_result.lower() and "processed" not in exec_result.lower():
            return (
                f"Step 5 failed: Batch OCR error.\n\n"
                f"{exec_result[:2000]}"
            )
        steps.append("OCR completed + combined workbook generated")

        # -- Step 6: Download URL --
        download_url = await _get_download_url(
            engine, session_id, output_filename
        )
        if download_url:
            steps.append("Download URL retrieved")
        else:
            steps.append("File created but URL not captured")

        # -- Report --
        report = [
            "## Batch OCR Pipeline \u2014 Complete",
            "",
            f"**Source:** OneDrive `{source_folder}`",
            f"**User:** {user_id}",
            f"**PDFs Processed:** {len(sandbox_paths)}",
            f"**Output:** `{output_filename}`",
        ]
        if download_url:
            report.append(f"**Download:** {download_url}")
        report.append("")
        report.append("### PDFs:")
        for sp in sandbox_paths:
            report.append(f"- {sp['name']}")
        report.append("")
        report.append("### Steps:")
        for s in steps:
            report.append(f"- {s}")
        report.append("")
        report.append("### Output:")
        report.append(exec_result[:2000] if exec_result else "(no output)")
        return "\n".join(report)

    except Exception as e:
        logger.error(f"skill_batch_ocr_pipeline: {e}", exc_info=True)
        lines = [
            "## Batch OCR Pipeline \u2014 Failed",
            f"**Error:** {e}",
            "",
            "### Steps completed:",
        ]
        for s in (steps or ["(none)"]):
            lines.append(f"- {s}")
        return "\n".join(lines)


SKILL_DEFINITION = {
    "name": "skill_batch_ocr_pipeline",
    "description": (
        "OCR multiple PDF files from an OneDrive folder into a single "
        "combined Excel workbook. Lists the folder, filters to PDFs, "
        "batch-downloads them, runs OCR on each, and combines all results "
        "into one workbook with a Summary sheet and per-PDF sheets. "
        "REQUIRES user_id (Microsoft 365 email). "
        "Pass source_folder (OneDrive path) and optionally max_files, "
        "file_filter, and output_filename."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "user_id": {
                "type": "string",
                "description": "Microsoft 365 email (REQUIRED)",
            },
            "source_folder": {
                "type": "string",
                "description": "OneDrive folder path containing PDFs",
                "default": "/",
            },
            "file_filter": {
                "type": "string",
                "description": "Extension filter (default '.pdf')",
                "default": ".pdf",
            },
            "output_filename": {
                "type": "string",
                "description": "Combined output Excel filename",
                "default": "batch_ocr_results.xlsx",
            },
            "max_files": {
                "type": "integer",
                "description": "Max PDFs to process (safety cap)",
                "default": 10,
            },
            "session_id": {
                "type": "string",
                "description": "Session ID (auto-created if empty)",
                "default": "",
            },
        },
        "required": ["user_id"],
    },
    "tools": ["onedrive", "execute_code", "list_files", "create_session"],
    "execute": execute_batch_ocr_pipeline,
}
