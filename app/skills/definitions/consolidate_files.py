"""Skill: Consolidate OneDrive Files to Excel

Downloads CSV/PDF/Excel files from a OneDrive folder and
consolidates them into a single Excel workbook with one tab
per source file.

This skill exists because the model consistently:
- Uses urllib instead of the onedrive tool
- Saves files to /tmp/ instead of the sandbox path
- Crashes on dataframe_to_rows
- Loops endlessly on file delivery
- Floods stdout with raw data

The skill enforces the correct behavior at the code level.
"""

import os
import re
import logging
from typing import Any, Dict, List, Optional

from ..base import Skill, SkillResult, StepResult, StepStatus
from ..engine import SkillContext
from ..guardrails import check_code_guardrails

logger = logging.getLogger(__name__)


# ── Code Templates ───────────────────────────────────────────
# These are the EXACT code blocks that get sent to execute_code.
# They use .format() for template vars and {{}} for runtime braces.
#
# Key decisions baked in:
#   - ws.append() instead of dataframe_to_rows (broken in sandbox)
#   - os.makedirs with exist_ok=True (prevent path errors)
#   - Print only summaries, never raw data
#   - Explicit output path in sandbox_data/{session}/

CSV_CONSOLIDATE_TEMPLATE = """import pandas as pd
from openpyxl import Workbook
import os

output_dir = "{output_dir}"
os.makedirs(output_dir, exist_ok=True)

wb = Workbook()
wb.remove(wb.active)

files = {file_list}
total_rows = 0

for file_path, tab_name in files:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.csv':
        df = pd.read_csv(file_path)
    elif ext in ('.xlsx', '.xls'):
        df = pd.read_excel(file_path)
    else:
        continue
    safe_name = tab_name[:31]
    ws = wb.create_sheet(title=safe_name)
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))
    total_rows += len(df)
    print(f"Tab '{{safe_name}}': {{len(df)}} rows")

output_path = os.path.join(output_dir, "{output_filename}")
wb.save(output_path)
size = os.path.getsize(output_path)
print(f"Saved: {{output_path}} ({{size}} bytes, {{len(wb.sheetnames)}} tabs, {{total_rows}} rows)")
"""

PDF_CONSOLIDATE_TEMPLATE = """import pdfplumber
from openpyxl import Workbook
import os

output_dir = "{output_dir}"
os.makedirs(output_dir, exist_ok=True)

wb = Workbook()
wb.remove(wb.active)

files = {file_list}
total_rows = 0

for file_path, tab_name in files:
    safe_name = tab_name[:31]
    ws = wb.create_sheet(title=safe_name)
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    clean = [str(c) if c else "" for c in row]
                    ws.append(clean)
                    total_rows += 1
    print(f"Tab '{{safe_name}}': extracted from PDF")

output_path = os.path.join(output_dir, "{output_filename}")
wb.save(output_path)
size = os.path.getsize(output_path)
print(f"Saved: {{output_path}} ({{size}} bytes, {{len(wb.sheetnames)}} tabs, {{total_rows}} rows)")
"""

VERIFY_FILE_TEMPLATE = """import os
path = "{output_dir}"
if os.path.isdir(path):
    for f in os.listdir(path):
        full = os.path.join(path, f)
        print(f"{{f}} — {{os.path.getsize(full)}} bytes")
else:
    print(f"Directory not found: {{path}}")
"""


class ConsolidateFilesSkill(Skill):
    """Download files from OneDrive and consolidate into Excel."""

    name = "skill_consolidate_files"
    description = (
        "Downloads CSV, PDF, or Excel files from a OneDrive folder "
        "and consolidates them into a single Excel workbook with "
        "one tab per source file. Handles authentication, download, "
        "processing, and file delivery automatically. Returns a "
        "download URL for the final workbook."
    )
    input_schema = {
        "type": "object",
        "properties": {
            "user_email": {
                "type": "string",
                "description": (
                    "Microsoft 365 email for OneDrive access "
                    "(e.g., user@bolthousefresh.com)"
                ),
            },
            "folder_path": {
                "type": "string",
                "description": (
                    "OneDrive folder path "
                    "(e.g., '/Reports/March 2026')"
                ),
            },
            "file_names": {
                "type": "array",
                "items": {"type": "string"},
                "description": (
                    "Specific file names to process. "
                    "If empty, processes all CSV/PDF/Excel files."
                ),
                "default": [],
            },
            "output_filename": {
                "type": "string",
                "description": "Output Excel file name",
                "default": "Consolidated_Output.xlsx",
            },
            "session_name": {
                "type": "string",
                "description": "Session name for sandbox storage",
                "default": "consolidate",
            },
        },
        "required": ["user_email", "folder_path"],
    }

    # Guardrails
    blocked_imports = [
        "urllib", "urllib.request", "requests",
        "http.client", "httpx", "aiohttp", "subprocess",
    ]
    max_code_lines = 40
    max_stdout_chars = 5000

    def validate_params(self, params: Dict[str, Any]) -> Optional[str]:
        if not params.get("user_email"):
            return "user_email is required"
        if not params.get("folder_path"):
            return "folder_path is required"
        if "@" not in params.get("user_email", ""):
            return "user_email must be a valid email address"
        return None

    async def execute(
        self, params: Dict[str, Any], ctx: SkillContext
    ) -> SkillResult:
        """Execute the full consolidation pipeline."""
        steps: List[StepResult] = []

        # Extract params with defaults
        user_email = params["user_email"]
        folder_path = params["folder_path"]
        file_names = params.get("file_names", [])
        output_filename = params.get(
            "output_filename", "Consolidated_Output.xlsx"
        )
        session_name = params.get("session_name", "consolidate")
        output_dir = f"/app/sandbox_data/{session_name}/"

        # ── Step 1: Authenticate ─────────────────────────
        step = await self._step_authenticate(user_email, ctx)
        steps.append(step)
        if step.status == StepStatus.FAILED:
            return SkillResult(
                skill_name=self.name,
                success=False,
                steps=steps,
                error=f"Auth failed: {step.error}",
            )

        # ── Step 2: List folder ──────────────────────────
        step, target_files = await self._step_list_folder(
            user_email, folder_path, file_names, ctx
        )
        steps.append(step)
        if step.status == StepStatus.FAILED:
            return SkillResult(
                skill_name=self.name,
                success=False,
                steps=steps,
                error=f"List failed: {step.error}",
            )

        # ── Step 3: Download files ───────────────────────
        step, downloaded = await self._step_download(
            user_email, folder_path, target_files, ctx
        )
        steps.append(step)
        if step.status == StepStatus.FAILED:
            return SkillResult(
                skill_name=self.name,
                success=False,
                steps=steps,
                error=f"Download failed: {step.error}",
            )

        # ── Step 4: Consolidate ──────────────────────────
        step = await self._step_consolidate(
            downloaded, output_dir, output_filename,
            session_name, ctx
        )
        steps.append(step)
        if step.status == StepStatus.FAILED:
            return SkillResult(
                skill_name=self.name,
                success=False,
                steps=steps,
                error=f"Processing failed: {step.error}",
            )

        # ── Step 5: Deliver ──────────────────────────────
        step = await self._step_deliver(
            session_name, output_dir, output_filename, ctx
        )
        steps.append(step)

        success = step.status == StepStatus.SUCCESS
        return SkillResult(
            skill_name=self.name,
            success=success,
            steps=steps,
            final_output=step.output if success else None,
            error=step.error if not success else None,
        )

    # ── Step Implementations ─────────────────────────────

    async def _step_authenticate(
        self, user_email: str, ctx: SkillContext
    ) -> StepResult:
        """Step 1: Ensure Microsoft auth is active."""
        try:
            result = await ctx.call_tool("ms_auth", {
                "action": "check",
                "user_email": user_email,
            })
            result_str = str(result).lower()

            if "authenticated" not in result_str and "valid" not in result_str:
                result = await ctx.call_tool("ms_auth", {
                    "action": "login",
                    "user_email": user_email,
                })

            return StepResult(
                step_name="authenticate",
                status=StepStatus.SUCCESS,
                output=f"Authenticated as {user_email}",
            )
        except Exception as e:
            return StepResult(
                step_name="authenticate",
                status=StepStatus.FAILED,
                error=str(e),
            )

    async def _step_list_folder(
        self,
        user_email: str,
        folder_path: str,
        file_names: List[str],
        ctx: SkillContext,
    ) -> tuple:
        """Step 2: List OneDrive folder and identify target files."""
        try:
            result = await ctx.call_tool("onedrive", {
                "action": "list",
                "path": folder_path,
                "user_email": user_email,
            })

            available = self._parse_file_list(result)

            # Filter to processable types
            ok_exts = {".csv", ".pdf", ".xlsx", ".xls"}
            targets = [
                f for f in available
                if os.path.splitext(f["name"])[1].lower() in ok_exts
            ]

            # Further filter if specific names requested
            if file_names:
                targets = [
                    f for f in targets
                    if f["name"] in file_names
                ]

            if not targets:
                return (
                    StepResult(
                        step_name="list_folder",
                        status=StepStatus.FAILED,
                        error=(
                            f"No processable files in {folder_path}. "
                            f"Found: {[f['name'] for f in available]}"
                        ),
                    ),
                    [],
                )

            names = [f["name"] for f in targets]
            return (
                StepResult(
                    step_name="list_folder",
                    status=StepStatus.SUCCESS,
                    output=f"Found {len(targets)} files: {names}",
                ),
                targets,
            )
        except Exception as e:
            return (
                StepResult(
                    step_name="list_folder",
                    status=StepStatus.FAILED,
                    error=str(e),
                ),
                [],
            )

    async def _step_download(
        self,
        user_email: str,
        folder_path: str,
        target_files: List[Dict],
        ctx: SkillContext,
    ) -> tuple:
        """Step 3: Download each file via the onedrive tool."""
        downloaded = []
        errors = []

        for f in target_files:
            try:
                file_path = f"{folder_path.rstrip('/')}/{f['name']}"
                result = await ctx.call_tool("onedrive", {
                    "action": "download",
                    "path": file_path,
                    "user_email": user_email,
                })

                local_path = self._parse_download_path(result)
                if local_path:
                    downloaded.append({
                        "name": f["name"],
                        "local_path": local_path,
                    })
                    logger.info(
                        f"[Skill] Downloaded: {f['name']} -> {local_path}"
                    )
                else:
                    errors.append(f"No path found for {f['name']}")
            except Exception as e:
                errors.append(f"{f['name']}: {e}")
                logger.warning(
                    f"[Skill] Download failed: {f['name']}: {e}"
                )

        if not downloaded:
            return (
                StepResult(
                    step_name="download_files",
                    status=StepStatus.FAILED,
                    error=f"All downloads failed: {errors}",
                ),
                [],
            )

        names = [d["name"] for d in downloaded]
        msg = f"Downloaded {len(downloaded)}/{len(target_files)}: {names}"
        if errors:
            msg += f" (errors: {errors})"

        return (
            StepResult(
                step_name="download_files",
                status=StepStatus.SUCCESS,
                output=msg,
            ),
            downloaded,
        )

    async def _step_consolidate(
        self,
        downloaded: List[Dict],
        output_dir: str,
        output_filename: str,
        session_name: str,
        ctx: SkillContext,
    ) -> StepResult:
        """Step 4: Build the Excel workbook via execute_code."""
        try:
            # Determine template
            has_pdf = any(
                d["name"].lower().endswith(".pdf")
                for d in downloaded
            )
            has_tabular = any(
                d["name"].lower().endswith((".csv", ".xlsx", ".xls"))
                for d in downloaded
            )

            file_list_str = repr([
                (d["local_path"], os.path.splitext(d["name"])[0])
                for d in downloaded
            ])

            if has_pdf and not has_tabular:
                template = PDF_CONSOLIDATE_TEMPLATE
            else:
                template = CSV_CONSOLIDATE_TEMPLATE

            code = template.format(
                output_dir=output_dir,
                file_list=file_list_str,
                output_filename=output_filename,
            )

            # Run guardrails check on generated code
            is_safe, violation = check_code_guardrails(
                code,
                blocked_imports=self.blocked_imports,
                max_lines=self.max_code_lines,
            )
            if not is_safe:
                return StepResult(
                    step_name="consolidate",
                    status=StepStatus.FAILED,
                    error=f"Code guardrail violated: {violation}",
                )

            result = await ctx.call_tool("execute_code", {
                "code": code,
                "session_name": session_name,
            })

            # Truncate output for the step result
            output_str = str(result)[:500]
            return StepResult(
                step_name="consolidate",
                status=StepStatus.SUCCESS,
                output=output_str,
            )

        except Exception as e:
            return StepResult(
                step_name="consolidate",
                status=StepStatus.FAILED,
                error=str(e),
            )

    async def _step_deliver(
        self,
        session_name: str,
        output_dir: str,
        output_filename: str,
        ctx: SkillContext,
    ) -> StepResult:
        """Step 5: Get the download URL (one retry max)."""
        try:
            # First attempt
            result = await ctx.call_tool("list_files", {
                "session_name": session_name,
            })
            url = self._parse_download_url(result, output_filename)

            if url:
                return StepResult(
                    step_name="deliver",
                    status=StepStatus.SUCCESS,
                    output=f"Download: {url}",
                )

            # Verify file exists before retry
            verify_code = VERIFY_FILE_TEMPLATE.format(
                output_dir=output_dir
            )
            await ctx.call_tool("execute_code", {
                "code": verify_code,
                "session_name": session_name,
            })

            # One retry
            result = await ctx.call_tool("list_files", {
                "session_name": session_name,
            })
            url = self._parse_download_url(result, output_filename)

            if url:
                return StepResult(
                    step_name="deliver",
                    status=StepStatus.SUCCESS,
                    output=f"Download: {url}",
                )

            # Give up with clear message
            return StepResult(
                step_name="deliver",
                status=StepStatus.FAILED,
                error=(
                    f"File exists at {output_dir}{output_filename} "
                    f"but list_files cannot generate a download URL. "
                    f"The file was built successfully."
                ),
            )

        except Exception as e:
            return StepResult(
                step_name="deliver",
                status=StepStatus.FAILED,
                error=str(e),
            )

    # ── Parsers ──────────────────────────────────────────

    @staticmethod
    def _parse_file_list(result: Any) -> List[Dict]:
        """Parse onedrive list result into file info dicts."""
        files = []
        if isinstance(result, dict):
            items = result.get("items", result.get("files", []))
            if isinstance(items, list):
                for item in items:
                    if isinstance(item, dict):
                        files.append({
                            "name": item.get("name", ""),
                            "size": item.get("size", 0),
                            "id": item.get("id", ""),
                        })

        # Fallback: parse from string
        if not files:
            result_str = str(result)
            names = re.findall(r"'name':\s*'([^']+)'", result_str)
            if not names:
                names = re.findall(r'"name":\s*"([^"]+)"', result_str)
            for name in names:
                files.append({"name": name, "size": 0, "id": ""})

        return files

    @staticmethod
    def _parse_download_path(result: Any) -> Optional[str]:
        """Extract local path from onedrive download result."""
        result_str = str(result)
        paths = re.findall(
            r"(/(?:app/sandbox_data|tmp)/[^\s'\"]+)", result_str
        )
        return paths[0] if paths else None

    @staticmethod
    def _parse_download_url(result: Any, filename: str) -> Optional[str]:
        """Extract download URL from list_files result."""
        result_str = str(result)
        urls = re.findall(r"(https?://[^\s'\"]+)", result_str)
        for url in urls:
            clean_fn = filename.replace(" ", "")
            clean_url = url.replace(" ", "")
            if clean_fn in clean_url or "/dl/" in url:
                return url
        return None
