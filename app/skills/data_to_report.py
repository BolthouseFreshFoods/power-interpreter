"""Skill: Data to Report — Automated data analysis + reporting.

Multi-step workflow that:
1. Loads a data file (CSV or Excel) from the sandbox
2. Validates columns exist
3. Computes aggregations grouped by a column
4. Generates a formatted Excel report with summary
5. Generates a matplotlib chart (auto-captured)
6. Returns download URLs for both files

Designed for Kevin's carrot harvest UPA reporting pattern.
Reduces 5-8 agent tool calls to 1 skill call.
"""

import logging
import re
import json
import asyncio

logger = logging.getLogger(__name__)

_DL_URL_PATTERN = (
    r'https://power-interpreter-production-6396'
    r'\.up\.railway\.app/dl/[a-f0-9-]+'
)

_UUID_PATTERN = (
    r'[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-'
    r'[0-9a-f]{4}-[0-9a-f]{12}'
)


async def _ensure_session(engine, session_id: str) -> str:
    """Ensure a session exists, creating one if needed."""
    if session_id and session_id.strip():
        return session_id.strip()
    result = await engine.call_tool("create_session")
    try:
        parsed = json.loads(result)
        sid = parsed.get("session_id", parsed.get("id", ""))
        if sid:
            return sid
    except (json.JSONDecodeError, AttributeError):
        pass
    match = re.search(_UUID_PATTERN, result)
    if match:
        return match.group()
    raise RuntimeError(f"Could not create session: {result[:200]}")


async def _get_download_url(engine, session_id: str, filename: str) -> str:
    """Get download URL with retry for list_files unreliability."""
    for attempt in range(3):
        files_result = await engine.call_tool("list_files", session_id=session_id)
        if filename in files_result:
            match = re.search(_DL_URL_PATTERN, files_result)
            if match:
                return match.group()
        if attempt < 2:
            await asyncio.sleep(1.5)
    return ""


async def execute_data_to_report(
    engine,
    input_file: str = "",
    report_title: str = "Data Report",
    group_by: str = "",
    value_columns: str = "",
    chart_type: str = "bar",
    output_filename: str = "report.xlsx",
    chart_filename: str = "report_chart.png",
    session_id: str = "",
) -> str:
    """Generate a formatted Excel report + chart from a data file.

    Args:
        engine: SkillEngine instance (injected)
        input_file: Path to CSV/Excel file in sandbox
        report_title: Title for the report header
        group_by: Column name to group data by
        value_columns: Comma-separated column names to aggregate
        chart_type: 'bar', 'line', or 'pie' (default 'bar')
        output_filename: Output Excel filename
        chart_filename: Output chart image filename
        session_id: Session ID (auto-created if empty)

    Returns:
        Status message with download URLs or error details
    """
    steps = []

    try:
        # -- Validate --
        if not input_file:
            return "Error: input_file is required (path to CSV or Excel in sandbox)."
        if not group_by:
            return "Error: group_by is required (column name to group data by)."
        if not value_columns:
            return "Error: value_columns is required (comma-separated column names)."
        if chart_type not in ("bar", "line", "pie"):
            return f"Error: chart_type must be 'bar', 'line', or 'pie'. Got '{chart_type}'."

        # -- Step 1: Session --
        session_id = await _ensure_session(engine, session_id)
        steps.append(f"Session ready ({session_id[:8]}...)")

        # -- Step 2: Generate report + chart --
        safe_input = input_file.replace("'", "\\'")
        safe_output = (
            "/app/sandbox_data/" + session_id + "/" + output_filename
        ).replace("'", "\\'")
        safe_chart = (
            "/app/sandbox_data/" + session_id + "/" + chart_filename
        ).replace("'", "\\'")
        safe_title = report_title.replace("'", "\\'")
        safe_group = group_by.replace("'", "\\'")
        safe_vals = value_columns.replace("'", "\\'")

        # Detect file format
        if safe_input.endswith('.csv') or safe_input.endswith('.tsv'):
            load_line = "df = pd.read_csv('" + safe_input + "')\n"
        else:
            load_line = "df = pd.read_excel('" + safe_input + "')\n"

        report_code = (
            "import pandas as pd\n"
            "import matplotlib\n"
            "matplotlib.use('Agg')\n"
            "import matplotlib.pyplot as plt\n"
            "from openpyxl import Workbook\n"
            "from openpyxl.styles import Font, PatternFill, Alignment, Border, Side\n"
            "\n"
            "# Load data\n"
            "print('Loading data...')\n"
            + load_line +
            "print(f'Loaded {len(df)} rows, {len(df.columns)} columns')\n"
            "print(f'Columns: {list(df.columns)}')\n"
            "\n"
            "# Validate columns\n"
            "group_col = '" + safe_group + "'\n"
            "val_cols = [c.strip() for c in '" + safe_vals + "'.split(',')]\n"
            "\n"
            "missing = []\n"
            "if group_col not in df.columns:\n"
            "    missing.append(f'group_by: {group_col}')\n"
            "for vc in val_cols:\n"
            "    if vc not in df.columns:\n"
            "        missing.append(f'value: {vc}')\n"
            "if missing:\n"
            "    print(f'ERROR: Missing columns: {missing}')\n"
            "    print(f'Available: {list(df.columns)}')\n"
            "    raise ValueError(f'Missing columns: {missing}')\n"
            "\n"
            "# Aggregate\n"
            "print(f'Grouping by {group_col}, aggregating {val_cols}...')\n"
            "numeric_vals = []\n"
            "for vc in val_cols:\n"
            "    df[vc] = pd.to_numeric(df[vc], errors='coerce')\n"
            "    numeric_vals.append(vc)\n"
            "\n"
            "summary = df.groupby(group_col)[numeric_vals].agg(['sum', 'mean', 'count'])\n"
            "summary_flat = summary.copy()\n"
            "summary_flat.columns = [f'{c[0]}_{c[1]}' for c in summary_flat.columns]\n"
            "summary_flat = summary_flat.reset_index()\n"
            "print(f'Summary: {len(summary_flat)} groups')\n"
            "\n"
            "# Excel workbook\n"
            "wb = Workbook()\n"
            "ws = wb.active\n"
            "ws.title = 'Summary'\n"
            "hf = Font(bold=True, color='FFFFFF')\n"
            "hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')\n"
            "tf = Font(bold=True, size=14)\n"
            "\n"
            "# Title\n"
            "ws.append(['" + safe_title + "'])\n"
            "ws['A1'].font = tf\n"
            "ws.append([])\n"
            "ws.append([f'Source: " + safe_input.split('/')[-1] + "'])\n"
            "ws.append([f'Rows: {len(df)}'])\n"
            "ws.append([f'Groups: {len(summary_flat)}'])\n"
            "ws.append([])\n"
            "\n"
            "# Header row\n"
            "headers = list(summary_flat.columns)\n"
            "ws.append(headers)\n"
            "for cell in ws[ws.max_row]:\n"
            "    cell.font = hf\n"
            "    cell.fill = hfill\n"
            "\n"
            "# Data rows\n"
            "for _, row in summary_flat.iterrows():\n"
            "    ws.append([round(v, 2) if isinstance(v, float) else v for v in row.values])\n"
            "\n"
            "# Auto-width columns\n"
            "for col in ws.columns:\n"
            "    mx = max(len(str(cell.value or '')) for cell in col)\n"
            "    ws.column_dimensions[col[0].column_letter].width = min(mx + 3, 40)\n"
            "\n"
            "# Raw data sheet\n"
            "ws2 = wb.create_sheet('Raw Data')\n"
            "ws2.append(list(df.columns))\n"
            "for cell in ws2[1]:\n"
            "    cell.font = hf\n"
            "    cell.fill = hfill\n"
            "for _, row in df.head(5000).iterrows():\n"
            "    ws2.append(list(row.values))\n"
            "\n"
            "wb.save('" + safe_output + "')\n"
            "print(f'\\nExcel saved to: " + safe_output + "')\n"
            "\n"
            "# Chart\n"
            "print('Generating chart...')\n"
            "fig, ax = plt.subplots(figsize=(12, 6))\n"
            "chart_data = df.groupby(group_col)[numeric_vals[0]].sum().sort_values(ascending=False)\n"
            "\n"
            "chart_type = '" + chart_type + "'\n"
            "if chart_type == 'bar':\n"
            "    chart_data.plot(kind='bar', ax=ax, color='#4472C4', edgecolor='white')\n"
            "    ax.set_ylabel(numeric_vals[0])\n"
            "elif chart_type == 'line':\n"
            "    chart_data.plot(kind='line', ax=ax, color='#4472C4', linewidth=2, marker='o')\n"
            "    ax.set_ylabel(numeric_vals[0])\n"
            "elif chart_type == 'pie':\n"
            "    chart_data.head(15).plot(kind='pie', ax=ax, autopct='%1.1f%%')\n"
            "    ax.set_ylabel('')\n"
            "\n"
            "ax.set_title('" + safe_title + "', fontsize=14, fontweight='bold')\n"
            "ax.set_xlabel(group_col)\n"
            "plt.xticks(rotation=45, ha='right')\n"
            "plt.tight_layout()\n"
            "plt.savefig('" + safe_chart + "', dpi=150, bbox_inches='tight',\n"
            "            facecolor='white', edgecolor='none')\n"
            "plt.close()\n"
            "print(f'Chart saved to: " + safe_chart + "')\n"
            "print(f'\\nReport complete: {len(summary_flat)} groups, {len(df)} rows')\n"
        )

        logger.info(
            f"skill_data_to_report: generating report from {input_file}"
        )
        exec_result = await engine.call_tool(
            "execute_code",
            session_id=session_id,
            code=report_code,
        )

        if "error" in exec_result.lower() and "saved" not in exec_result.lower() and "complete" not in exec_result.lower():
            return (
                f"Step 2 failed: Report generation error.\n\n"
                f"{exec_result[:2000]}"
            )
        steps.append("Excel report + chart generated")

        # -- Step 3: Download URLs --
        excel_url = await _get_download_url(
            engine, session_id, output_filename
        )
        chart_url = await _get_download_url(
            engine, session_id, chart_filename
        )

        if excel_url:
            steps.append("Excel download URL retrieved")
        if chart_url:
            steps.append("Chart download URL retrieved")
        if not excel_url and not chart_url:
            steps.append("Files created but URLs not captured")

        # -- Report --
        report = [
            "## Data Report \u2014 Complete",
            "",
            f"**Title:** {report_title}",
            f"**Source:** `{input_file}`",
            f"**Group By:** {group_by}",
            f"**Values:** {value_columns}",
            f"**Chart Type:** {chart_type}",
        ]
        if excel_url:
            report.append(f"**Excel Download:** {excel_url}")
        if chart_url:
            report.append(f"**Chart Download:** {chart_url}")
        report.append("")
        report.append("### Steps:")
        for s in steps:
            report.append(f"- {s}")
        report.append("")
        report.append("### Output:")
        report.append(exec_result[:2000] if exec_result else "(no output)")
        return "\n".join(report)

    except Exception as e:
        logger.error(f"skill_data_to_report: {e}", exc_info=True)
        lines = [
            "## Data Report \u2014 Failed",
            f"**Error:** {e}",
            "",
            "### Steps completed:",
        ]
        for s in (steps or ["(none)"]):
            lines.append(f"- {s}")
        return "\n".join(lines)


SKILL_DEFINITION = {
    "name": "skill_data_to_report",
    "description": (
        "Generate a formatted Excel report and summary chart from a data file. "
        "Loads CSV or Excel, groups by a column, aggregates numeric values, "
        "and produces a professional Excel workbook (Summary + Raw Data sheets) "
        "plus a matplotlib chart. Returns download URLs for both files. "
        "Pass input_file (sandbox path), group_by (column name), and "
        "value_columns (comma-separated). Optionally set chart_type and report_title."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "input_file": {
                "type": "string",
                "description": "Path to CSV/Excel file in sandbox",
            },
            "report_title": {
                "type": "string",
                "description": "Report title",
                "default": "Data Report",
            },
            "group_by": {
                "type": "string",
                "description": "Column name to group data by",
            },
            "value_columns": {
                "type": "string",
                "description": "Comma-separated columns to aggregate",
            },
            "chart_type": {
                "type": "string",
                "description": "'bar', 'line', or 'pie'",
                "default": "bar",
                "enum": ["bar", "line", "pie"],
            },
            "output_filename": {
                "type": "string",
                "description": "Output Excel filename",
                "default": "report.xlsx",
            },
            "chart_filename": {
                "type": "string",
                "description": "Output chart filename",
                "default": "report_chart.png",
            },
            "session_id": {
                "type": "string",
                "description": "Session ID (auto-created if empty)",
                "default": "",
            },
        },
        "required": ["input_file", "group_by", "value_columns"],
    },
    "tools": ["execute_code", "list_files", "create_session"],
    "execute": execute_data_to_report,
}
